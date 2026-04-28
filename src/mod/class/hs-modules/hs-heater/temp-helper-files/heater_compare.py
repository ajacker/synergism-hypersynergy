from pathlib import Path
import json
import openpyxl
import argparse

WORKBOOK_RELATIVE = Path("../../../../../../docs/SynergismOfficial-master/Re-Octptimisation (Beta 1.2.5E)/Re-Octptimisation (Beta 1.2.5E).xlsx")
SHEET_NAME = "True Ambrosia Heater"
LABEL_TO_FIELD = {
    'Asc. Speed (C)': 'ascSpeed',
    'Ascension Speed (H)': 'ascSpeed2',
    'SI Non-Amb RC': 'sirc',
    'Bonus SI': 'bonussi',
    'Bonus IA (total)': 'totalbonusia',
    'Bonus IA (from talisman)': 'talismanbonusia',
    'Base Talisman Power': 'baseTalismanPower',
    'Rune Exp (log) (>50)': 'runeexp',
    'Blueberries': 'blueberries',
    'Base Obtainium': 'baseobt',
    'Base Offering': 'baseoff',
    'Voucher': 'voucher',
}


def read_sheet_values(workbook_path: Path):
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}")

    sheet = wb[SHEET_NAME]
    found = {}
    for row in sheet.iter_rows(values_only=True):
        label = row[0]
        if label is None:
            continue
        label = str(label).strip()
        if label in LABEL_TO_FIELD:
            field = LABEL_TO_FIELD[label]
            found[field] = {
                'label': label,
                'value': row[1] if len(row) > 1 else None,
                'ref': row[4] if len(row) > 4 else None,
            }
    return found


def compare_export(sample_path: Path, sheet_values: dict):
    if not sample_path.exists():
        raise FileNotFoundError(f"Sample export not found: {sample_path}")
    sample = json.loads(sample_path.read_text(encoding='utf-8'))
    if 'hs_data' not in sample:
        raise ValueError('Sample JSON missing hs_data root object')
    export = sample['hs_data']

    results = []
    for field, metadata in sheet_values.items():
        actual = export.get(field)
        results.append({
            'field': field,
            'label': metadata['label'],
            'sheet_value': metadata['value'],
            'export_value': actual,
            'match': actual == metadata['value'],
        })
    return results


def main():
    parser = argparse.ArgumentParser(description='Compare Ambrosia Heater export fields against the True Ambrosia Heater spreadsheet.')
    parser.add_argument('--sample', type=Path, help='Path to a JSON heater export sample file')
    args = parser.parse_args()

    workbook_path = (Path(__file__).resolve().parent / WORKBOOK_RELATIVE).resolve()
    sheet_values = read_sheet_values(workbook_path)

    print('Spreadsheet values from', workbook_path)
    for field, metadata in sheet_values.items():
        print(f"{field}: label={metadata['label']} value={metadata['value']} ref={metadata['ref']}")

    if args.sample:
        print('\nComparing to sample export', args.sample)
        results = compare_export(args.sample, sheet_values)
        for result in results:
            status = 'MATCH' if result['match'] else 'MISMATCH'
            print(f"{status}: {result['field']} sheet={result['sheet_value']} export={result['export_value']}")


if __name__ == '__main__':
    main()
